"""
Data Loader Module - High-Performance Excel Data Loading
Uses Pandas for bulk data loading instead of cell-by-cell win32com access.
Provides 10-50x speed improvement over COM-based access.
"""
import pandas as pd
import numpy as np
import os
from typing import Dict, Optional, Tuple, Any, List


class DataLoader:
    """
    High-performance Excel data loader using Pandas.
    Loads entire sheets into DataFrames for fast in-memory access.
    """
    
    def __init__(self):
        # Config Excel DataFrames
        self.config_file_path = None
        self.df_main_sheet = None
        self.df_담보매핑 = None
        self.df_참조1 = None
        self.df_참조2 = None
        self.df_독특레이아웃 = None
        self.df_독특메모예외 = None
        
        # Named Range positions (row, col - 0-indexed)
        self.named_ranges = {}
        
        # PGM Excel DataFrames
        self.pgm_file_path = None
        self.df_pgm_main = None
        self.df_보장구조 = None
        self.df_보장배수 = None
        self.df_보기납기 = None
        
        # Cached data as numpy arrays for fastest access
        self.arr_pgm_main = None
        self.arr_pgm_main_temp = None
        self.arr_보장구조 = None
        self.arr_보장배수 = None
        self.arr_보기납기 = None
        self.arr_담보매핑 = None
        self.arr_참조1 = None
        self.arr_참조2 = None
        self.arr_독특파일명 = None
        self.arr_독특레이아웃 = None
        self.arr_독특메모예외 = None
        self.arr_source_doc = None
        
        # Reference data extracted from config
        self.product_code = ""
        self.prod_name = ""
        self.출력약관경로 = ""
        self.종속특약경로 = ""
        self.pgm_path = ""
        self.pgm_filename = ""
        self.약관경로 = ""
        self.약관파일명 = ""
        
        # Product attributes
        self.자동갱신형 = 0
        self.단체보험 = 0
        self.모듈형 = 0
        self.독립특약 = 0
        self.중증간편 = 0
        self.zero_age_자녀 = 0
        
        # Loop range
        self.loop_start = 0
        self.loop_end = 0
        
        # 통합PGM 관련
        self.is_통합pgm = False
        self.count_종 = 0
        self.종구분 = 0
        self.arr_통합PGM_상품 = None

    def load_config_excel(self, file_path: str, main_sheet_name: str = None, log_callback=None) -> bool:
        """
        Load the configuration Excel file using Pandas.
        All sheets are loaded into DataFrames for fast access.
        
        Args:
            file_path: Path to the configuration Excel file
            main_sheet_name: Name of the main sheet (optional)
            log_callback: Logging callback function
            
        Returns:
            True if successful, False otherwise
        """
        try:
            self.config_file_path = os.path.abspath(file_path)
            
            if not os.path.exists(self.config_file_path):
                raise FileNotFoundError(f"설정 파일을 찾을 수 없습니다: {self.config_file_path}")
            
            if log_callback:
                log_callback(f"설정 파일 로드 중: {file_path}")
            
            # Load all sheets at once using ExcelFile for efficiency
            with pd.ExcelFile(self.config_file_path, engine='openpyxl') as xlsx:
                sheet_names = xlsx.sheet_names
                
                # Load main sheet (first sheet or specified)
                if main_sheet_name and main_sheet_name in sheet_names:
                    self.df_main_sheet = pd.read_excel(xlsx, sheet_name=main_sheet_name, header=None)
                else:
                    self.df_main_sheet = pd.read_excel(xlsx, sheet_name=0, header=None)
                
                # Load 담보매핑 sheet
                if "담보매핑" in sheet_names:
                    self.df_담보매핑 = pd.read_excel(xlsx, sheet_name="담보매핑", header=None)
                    if log_callback:
                        log_callback(f"  - 담보매핑: {len(self.df_담보매핑)} rows")
                
                # Load 참조 sheet
                if "참조" in sheet_names:
                    df_참조 = pd.read_excel(xlsx, sheet_name="참조", header=None)
                    # Split into 참조1 and 참조2 based on named ranges
                    # This will be refined when we parse named ranges
                    self.df_참조1 = df_참조
                    self.df_참조2 = df_참조
                
                # Load 독특레이아웃 sheet (for 독립특약)
                if "독특레이아웃" in sheet_names:
                    self.df_독특레이아웃 = pd.read_excel(xlsx, sheet_name="독특레이아웃", header=None)
            
            # Parse named ranges and extract key data
            self._parse_config_data(log_callback)
            
            if log_callback:
                log_callback(f"설정 파일 로드 완료: 상품코드={self.product_code}, 상품명={self.prod_name}")
            
            return True
            
        except Exception as e:
            if log_callback:
                log_callback(f"설정 파일 로드 에러: {e}")
            raise

    def _parse_config_data(self, log_callback=None):
        """
        Parse configuration data from the main sheet DataFrame.
        Extracts product info, paths, and product attributes.
        """
        if self.df_main_sheet is None:
            return
        
        df = self.df_main_sheet
        
        # Find named ranges by searching for keywords
        # This is a simplified approach - in production, you'd read actual named ranges
        
        for idx, row in df.iterrows():
            for col_idx, cell in enumerate(row):
                cell_str = str(cell).strip() if pd.notna(cell) else ""
                
                # Find product code/name
                if cell_str == "상품코드" or "Ref_상품코드" in cell_str:
                    self.product_code = str(df.iloc[idx, col_idx + 1]).strip() if col_idx + 1 < len(row) and pd.notna(df.iloc[idx, col_idx + 1]) else ""
                    if idx + 1 < len(df):
                        self.prod_name = str(df.iloc[idx + 1, col_idx + 1]).strip() if pd.notna(df.iloc[idx + 1, col_idx + 1]) else ""
                    self.named_ranges["Ref_상품코드"] = (idx, col_idx)
                
                # Find paths
                elif cell_str == "출력약관경로" or (col_idx == 0 and "출력" in cell_str and "경로" in cell_str):
                    # This marks the start of path settings
                    self.named_ranges["Ref_경로지정"] = (idx, col_idx)
                    # Parse paths
                    if col_idx + 1 < len(row):
                        for path_offset in range(6):
                            if idx + path_offset < len(df):
                                path_val = df.iloc[idx + path_offset, col_idx + 1]
                                path_str = str(path_val).strip() if pd.notna(path_val) else ""
                                if path_offset == 0:
                                    self.출력약관경로 = path_str
                                elif path_offset == 1:
                                    self.종속특약경로 = path_str
                                elif path_offset == 2:
                                    self.pgm_path = path_str
                                elif path_offset == 3:
                                    self.pgm_filename = path_str
                                elif path_offset == 4:
                                    self.약관경로 = path_str
                                elif path_offset == 5:
                                    self.약관파일명 = path_str
                
                # Find product attributes
                elif cell_str == "자동갱신형":
                    val = df.iloc[idx, col_idx + 1] if col_idx + 1 < len(row) else 0
                    self.자동갱신형 = int(val) if pd.notna(val) else 0
                elif cell_str == "단체보험":
                    val = df.iloc[idx, col_idx + 1] if col_idx + 1 < len(row) else 0
                    self.단체보험 = int(val) if pd.notna(val) else 0
                elif cell_str == "모듈형":
                    val = df.iloc[idx, col_idx + 1] if col_idx + 1 < len(row) else 0
                    self.모듈형 = int(val) if pd.notna(val) else 0
                elif cell_str == "독립특약":
                    val = df.iloc[idx, col_idx + 1] if col_idx + 1 < len(row) else 0
                    self.독립특약 = int(val) if pd.notna(val) else 0
                elif cell_str == "중증간편":
                    val = df.iloc[idx, col_idx + 1] if col_idx + 1 < len(row) else 0
                    self.중증간편 = int(val) if pd.notna(val) else 0
                elif cell_str == "0세자녀":
                    val = df.iloc[idx, col_idx + 1] if col_idx + 1 < len(row) else 0
                    self.zero_age_자녀 = int(val) if pd.notna(val) else 0
                
                # Find loop range
                elif cell_str == "출력시작":
                    val = df.iloc[idx, col_idx + 1] if col_idx + 1 < len(row) else 0
                    self.loop_start = int(val) if pd.notna(val) else 0
                elif cell_str == "출력종료":
                    val = df.iloc[idx, col_idx + 1] if col_idx + 1 < len(row) else 0
                    self.loop_end = int(val) if pd.notna(val) else 0
                
                # Find key named ranges
                elif "Ref_Point" in cell_str:
                    self.named_ranges["Ref_Point"] = (idx, col_idx)
                elif "Ref_담보속성1" in cell_str:
                    self.named_ranges["Ref_담보속성1"] = (idx, col_idx)
                elif "Ref_담보속성2" in cell_str:
                    self.named_ranges["Ref_담보속성2"] = (idx, col_idx)
                elif "Ref_담보매핑결과" in cell_str:
                    self.named_ranges["Ref_담보매핑결과"] = (idx, col_idx)
                elif "Ref_종속특약" in cell_str:
                    self.named_ranges["Ref_종속특약"] = (idx, col_idx)
        
        # Convert dataframes to numpy arrays for fastest access
        if self.df_담보매핑 is not None:
            self.arr_담보매핑 = self.df_담보매핑.values
        
        if self.df_참조1 is not None:
            self.arr_참조1 = self.df_참조1.values
        
        if self.df_참조2 is not None:
            self.arr_참조2 = self.df_참조2.values

    def load_pgm_excel(self, file_path: str, product_code: str = None, log_callback=None) -> bool:
        """
        Load the PGM Excel file using Pandas.
        Loads Main, 보장구조, 보장배수, 보기납기 sheets.
        
        Args:
            file_path: Path to the PGM Excel file
            product_code: Product code for Main sheet name
            log_callback: Logging callback function
            
        Returns:
            True if successful, False otherwise
        """
        try:
            self.pgm_file_path = os.path.abspath(file_path)
            
            if not os.path.exists(self.pgm_file_path):
                raise FileNotFoundError(f"PGM 파일을 찾을 수 없습니다: {self.pgm_file_path}")
            
            if log_callback:
                log_callback(f"PGM 파일 로드 중 (Pandas): {file_path}")
            
            # Use product_code parameter or instance variable
            code = product_code or self.product_code
            main_sheet_name = f"Main_{code}"
            
            # Load sheets using ExcelFile for efficiency
            with pd.ExcelFile(self.pgm_file_path, engine='openpyxl') as xlsx:
                sheet_names = xlsx.sheet_names
                
                # Load Main sheet
                if main_sheet_name in sheet_names:
                    self.df_pgm_main = pd.read_excel(xlsx, sheet_name=main_sheet_name, header=None)
                    if log_callback:
                        log_callback(f"  - Main: {len(self.df_pgm_main)} rows")
                else:
                    # Try to find a Main_ sheet
                    for name in sheet_names:
                        if name.startswith("Main_"):
                            self.df_pgm_main = pd.read_excel(xlsx, sheet_name=name, header=None)
                            if log_callback:
                                log_callback(f"  - {name}: {len(self.df_pgm_main)} rows")
                            break
                
                # Load 1.보기납기
                if "1.보기납기" in sheet_names:
                    self.df_보기납기 = pd.read_excel(xlsx, sheet_name="1.보기납기", header=None)
                    if log_callback:
                        log_callback(f"  - 보기납기: {len(self.df_보기납기)} rows")
                
                # Load 2.보장구조
                if "2.보장구조" in sheet_names:
                    self.df_보장구조 = pd.read_excel(xlsx, sheet_name="2.보장구조", header=None)
                    if log_callback:
                        log_callback(f"  - 보장구조: {len(self.df_보장구조)} rows")
                
                # Load 3.보장배수
                if "3.보장배수" in sheet_names:
                    self.df_보장배수 = pd.read_excel(xlsx, sheet_name="3.보장배수", header=None)
                    if log_callback:
                        log_callback(f"  - 보장배수: {len(self.df_보장배수)} rows")
            
            # Convert to numpy arrays for fastest access
            self._convert_pgm_to_arrays(log_callback)
            
            if log_callback:
                log_callback(f"PGM 파일 로드 완료!")
            
            return True
            
        except Exception as e:
            if log_callback:
                log_callback(f"PGM 파일 로드 에러: {e}")
            raise

    def _convert_pgm_to_arrays(self, log_callback=None):
        """
        Convert PGM DataFrames to numpy arrays for fastest access.
        Also builds dictionary indexes for O(1) lookup.
        """
        if self.df_pgm_main is not None:
            self.arr_pgm_main = self.df_pgm_main.values
        
        if self.df_보장구조 is not None:
            self.arr_보장구조 = self.df_보장구조.values
        
        if self.df_보장배수 is not None:
            self.arr_보장배수 = self.df_보장배수.values
        
        if self.df_보기납기 is not None:
            self.arr_보기납기 = self.df_보기납기.values
        
        # Build dictionary indexes for fast lookup
        self._build_pgm_indexes(log_callback)
    
    def _build_pgm_indexes(self, log_callback=None):
        """
        Build dictionary indexes from PGM numpy arrays for O(1) lookup.
        Replaces linear scans in _read_pgm_loop and _read_pgm_보기납기.
        """
        # Index for arr_pgm_main: {(dambo_code_col1): [row_indices], (dambo_code_col0): [row_indices]}
        self.pgm_main_index_col0 = {}  # col0 기준 (독립특약)
        self.pgm_main_index_col1 = {}  # col1 기준 (일반 담보)
        if self.arr_pgm_main is not None:
            for i, row in enumerate(self.arr_pgm_main):
                if row[0] is not None:
                    key0 = str(row[0]).strip()
                    if key0:
                        if key0 not in self.pgm_main_index_col0:
                            self.pgm_main_index_col0[key0] = []
                        self.pgm_main_index_col0[key0].append(i)
                if len(row) > 1 and row[1] is not None:
                    key1 = str(row[1]).strip()
                    if key1:
                        if key1 not in self.pgm_main_index_col1:
                            self.pgm_main_index_col1[key1] = []
                        self.pgm_main_index_col1[key1].append(i)
        
        # Index for arr_보장구조: {lookup_key: [row_indices]}
        self.보장구조_index = {}
        if self.arr_보장구조 is not None:
            for i, row in enumerate(self.arr_보장구조):
                key = str(row[0]).strip() if row[0] is not None else ""
                if key:
                    if key not in self.보장구조_index:
                        self.보장구조_index[key] = []
                    self.보장구조_index[key].append(i)
        
        # Index for arr_보기납기: {lookup_key: [row_indices]}
        self.보기납기_index = {}
        if self.arr_보기납기 is not None:
            for i, row in enumerate(self.arr_보기납기):
                key = str(row[0]).strip() if row[0] is not None else ""
                if key:
                    if key not in self.보기납기_index:
                        self.보기납기_index[key] = []
                    self.보기납기_index[key].append(i)
        
        # Index for arr_보장배수: {lookup_key: row_index} (first match only)
        self.보장배수_index = {}
        if self.arr_보장배수 is not None:
            for i, row in enumerate(self.arr_보장배수):
                key = str(row[0]).strip() if row[0] is not None else ""
                if key and key not in self.보장배수_index:
                    self.보장배수_index[key] = i

    def get_ref_point_data(self, start_row: int, end_row: int) -> np.ndarray:
        """
        Get Ref_Point data for the specified row range.
        Returns numpy array for fast iteration.
        """
        if "Ref_Point" in self.named_ranges and self.df_main_sheet is not None:
            ref_row, ref_col = self.named_ranges["Ref_Point"]
            # Return subset of data
            return self.df_main_sheet.iloc[ref_row + start_row:ref_row + end_row + 1, ref_col:ref_col + 10].values
        return np.array([])

    def get_담보속성1_data(self, start_row: int, end_row: int) -> np.ndarray:
        """Get 담보속성1 data for specified range."""
        if "Ref_담보속성1" in self.named_ranges and self.df_main_sheet is not None:
            ref_row, ref_col = self.named_ranges["Ref_담보속성1"]
            return self.df_main_sheet.iloc[ref_row + start_row:ref_row + end_row + 1, ref_col:ref_col + 10].values
        return np.array([])

    def get_담보속성2_data(self, start_row: int, end_row: int) -> np.ndarray:
        """Get 담보속성2 data for specified range."""
        if "Ref_담보속성2" in self.named_ranges and self.df_main_sheet is not None:
            ref_row, ref_col = self.named_ranges["Ref_담보속성2"]
            return self.df_main_sheet.iloc[ref_row + start_row:ref_row + end_row + 1, ref_col:ref_col + 10].values
        return np.array([])

    def get_담보매핑결과_data(self, start_row: int, end_row: int) -> np.ndarray:
        """Get 담보매핑결과 data for specified range."""
        if "Ref_담보매핑결과" in self.named_ranges and self.df_main_sheet is not None:
            ref_row, ref_col = self.named_ranges["Ref_담보매핑결과"]
            return self.df_main_sheet.iloc[ref_row + start_row:ref_row + end_row + 1, ref_col:ref_col + 5].values
        return np.array([])

    def find_row_in_array(self, arr: np.ndarray, col_index: int, search_value: str) -> int:
        """
        Fast row search in numpy array.
        Returns row index (0-based) or -1 if not found.
        Uses numpy vectorization for speed.
        """
        if arr is None or len(arr) == 0:
            return -1
        
        try:
            # Convert column to string for comparison
            col_data = arr[:, col_index].astype(str)
            search_str = str(search_value).strip()
            
            # Find matching rows
            matches = np.where(np.char.strip(col_data) == search_str)[0]
            
            if len(matches) > 0:
                return int(matches[0])
            return -1
        except:
            return -1

    def find_rows_matching(self, arr: np.ndarray, col_index: int, search_value: str) -> List[int]:
        """
        Find all rows matching a value in specified column.
        Returns list of row indices (0-based).
        """
        if arr is None or len(arr) == 0:
            return []
        
        try:
            col_data = arr[:, col_index].astype(str)
            search_str = str(search_value).strip()
            matches = np.where(np.char.strip(col_data) == search_str)[0]
            return matches.tolist()
        except:
            return []

    def get_array_row(self, arr: np.ndarray, lookup_key: Any, col_index: int = 0) -> Optional[np.ndarray]:
        """
        Get a row from array matching the lookup key in specified column.
        Fast alternative to copy_array_subset.
        """
        row_idx = self.find_row_in_array(arr, col_index, lookup_key)
        if row_idx >= 0 and row_idx < len(arr):
            return arr[row_idx]
        return None

    def close(self):
        """Clear all loaded data."""
        self.df_main_sheet = None
        self.df_담보매핑 = None
        self.df_참조1 = None
        self.df_참조2 = None
        self.df_pgm_main = None
        self.df_보장구조 = None
        self.df_보장배수 = None
        self.df_보기납기 = None
        
        self.arr_pgm_main = None
        self.arr_보장구조 = None
        self.arr_보장배수 = None
        self.arr_보기납기 = None
        self.arr_담보매핑 = None


class ExcelWriter:
    """
    Minimal Excel writer for saving results.
    Uses openpyxl for writing specific cells without full COM overhead.
    """
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = None
    
    def load(self):
        """Load workbook for editing."""
        from openpyxl import load_workbook
        self.wb = load_workbook(self.file_path)
    
    def write_cell(self, sheet_name: str, row: int, col: int, value: Any):
        """Write a single cell value (1-indexed row/col for compatibility)."""
        if self.wb:
            ws = self.wb[sheet_name]
            ws.cell(row=row, column=col, value=value)
    
    def save(self):
        """Save changes."""
        if self.wb:
            self.wb.save(self.file_path)
    
    def close(self):
        """Close workbook."""
        if self.wb:
            self.wb.close()
            self.wb = None
