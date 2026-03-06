import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .accuracy import AccuracyMetrics

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generate comparison report Excel with multiple sheets."""

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def generate(
        self,
        method1_metrics: AccuracyMetrics,
        method2_metrics: AccuracyMetrics,
        method1_timing: Optional[dict] = None,
        method2_timing: Optional[dict] = None,
        output_dir: str = ".",
    ) -> str:
        """Generate comparison_report.xlsx with Summary, Field_Accuracy, Mismatches, Timing sheets."""
        output_path = Path(output_dir) / "comparison_report.xlsx"
        wb = Workbook()

        # Sheet 1: Summary
        self._write_summary(wb.active, method1_metrics, method2_metrics)

        # Sheet 2: Field Accuracy
        ws_field = wb.create_sheet("Field_Accuracy")
        self._write_field_accuracy(ws_field, method1_metrics, method2_metrics)

        # Sheet 3: Mismatches
        ws_mis = wb.create_sheet("Mismatches")
        self._write_mismatches(ws_mis, method1_metrics, method2_metrics)

        # Sheet 4: Timing
        ws_time = wb.create_sheet("Timing")
        self._write_timing(ws_time, method1_timing or {}, method2_timing or {})

        wb.save(str(output_path))
        logger.info(f"Report saved to {output_path}")
        return str(output_path)

    def _style_header_row(self, ws, row: int, cols: int):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.THIN_BORDER

    def _write_summary(self, ws, m1: AccuracyMetrics, m2: AccuracyMetrics):
        ws.title = "Summary"
        headers = ["지표", "Method 1 (Baseline)", "Method 2 (Adobe)", "승자"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header_row(ws, 1, len(headers))

        rows_data = [
            ("정답 행 수", m1.total_rows_expected, m2.total_rows_expected, "-"),
            ("추출 행 수", m1.total_rows_extracted, m2.total_rows_extracted, "-"),
            ("담보 수", m1.total_coverages, m2.total_coverages, "-"),
            (
                "행 일치 수 (집합 교집합)",
                m1.total_rows_matched,
                m2.total_rows_matched,
                self._winner(m1.total_rows_matched, m2.total_rows_matched),
            ),
            (
                "행 일치율",
                f"{m1.row_match_rate:.1%}",
                f"{m2.row_match_rate:.1%}",
                self._winner(m1.row_match_rate, m2.row_match_rate),
            ),
            (
                "담보 완전일치 수",
                m1.perfect_coverages,
                m2.perfect_coverages,
                self._winner(m1.perfect_coverages, m2.perfect_coverages),
            ),
            (
                "담보 완전일치율",
                f"{m1.coverage_perfect_rate:.1%}",
                f"{m2.coverage_perfect_rate:.1%}",
                self._winner(m1.coverage_perfect_rate, m2.coverage_perfect_rate),
            ),
            (
                "담보 커버리지",
                f"{m1.coverage_completeness:.1%}",
                f"{m2.coverage_completeness:.1%}",
                self._winner(m1.coverage_completeness, m2.coverage_completeness),
            ),
            ("누락 행 수", len(m1.missing_rows), len(m2.missing_rows),
             self._winner(len(m2.missing_rows), len(m1.missing_rows))),
            ("초과 행 수", len(m1.extra_rows), len(m2.extra_rows),
             self._winner(len(m2.extra_rows), len(m1.extra_rows))),
        ]

        for r_idx, (label, v1, v2, winner) in enumerate(rows_data, 2):
            ws.cell(row=r_idx, column=1, value=label).border = self.THIN_BORDER
            ws.cell(row=r_idx, column=2, value=v1).border = self.THIN_BORDER
            ws.cell(row=r_idx, column=3, value=v2).border = self.THIN_BORDER
            cell_w = ws.cell(row=r_idx, column=4, value=winner)
            cell_w.border = self.THIN_BORDER
            if winner == "Method 1":
                cell_w.fill = self.GOOD_FILL
            elif winner == "Method 2":
                cell_w.fill = self.GOOD_FILL

        # Auto-fit column widths
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 22

    def _write_field_accuracy(self, ws, m1: AccuracyMetrics, m2: AccuracyMetrics):
        headers = ["필드", "Method 1 (%)", "Method 2 (%)", "승자"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header_row(ws, 1, len(headers))

        all_fields = sorted(
            set(list(m1.field_accuracy.keys()) + list(m2.field_accuracy.keys()))
        )
        for r_idx, field_name in enumerate(all_fields, 2):
            v1 = m1.field_accuracy.get(field_name, 0.0)
            v2 = m2.field_accuracy.get(field_name, 0.0)
            ws.cell(row=r_idx, column=1, value=field_name).border = self.THIN_BORDER
            ws.cell(row=r_idx, column=2, value=f"{v1:.1%}").border = self.THIN_BORDER
            ws.cell(row=r_idx, column=3, value=f"{v2:.1%}").border = self.THIN_BORDER
            ws.cell(
                row=r_idx, column=4, value=self._winner(v1, v2)
            ).border = self.THIN_BORDER

        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 20

    def _write_mismatches(self, ws, m1: AccuracyMetrics, m2: AccuracyMetrics):
        headers = ["Method", "담보명", "필드", "추출값", "정답값"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header_row(ws, 1, len(headers))

        row = 2
        for mm in m1.field_mismatches:
            for field_name, diff in mm.get("differences", {}).items():
                ws.cell(row=row, column=1, value="Method 1")
                ws.cell(row=row, column=2, value=mm.get("담보명", ""))
                ws.cell(row=row, column=3, value=field_name)
                ws.cell(row=row, column=4, value=diff.get("extracted", ""))
                ws.cell(row=row, column=5, value=diff.get("expected", ""))
                for c in range(1, 6):
                    ws.cell(row=row, column=c).border = self.THIN_BORDER
                row += 1

        for mm in m2.field_mismatches:
            for field_name, diff in mm.get("differences", {}).items():
                ws.cell(row=row, column=1, value="Method 2")
                ws.cell(row=row, column=2, value=mm.get("담보명", ""))
                ws.cell(row=row, column=3, value=field_name)
                ws.cell(row=row, column=4, value=diff.get("extracted", ""))
                ws.cell(row=row, column=5, value=diff.get("expected", ""))
                for c in range(1, 6):
                    ws.cell(row=row, column=c).border = self.THIN_BORDER
                row += 1

        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 22

    def _write_timing(self, ws, t1: dict, t2: dict):
        headers = ["단계", "Method 1 (초)", "Method 2 (초)"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        self._style_header_row(ws, 1, len(headers))

        stages = ["conversion", "table_parsing", "llm_extraction", "total"]
        for r_idx, stage in enumerate(stages, 2):
            ws.cell(row=r_idx, column=1, value=stage).border = self.THIN_BORDER
            ws.cell(
                row=r_idx, column=2, value=f"{t1.get(stage, 0):.2f}"
            ).border = self.THIN_BORDER
            ws.cell(
                row=r_idx, column=3, value=f"{t2.get(stage, 0):.2f}"
            ).border = self.THIN_BORDER

        for col in range(1, 4):
            ws.column_dimensions[chr(64 + col)].width = 20

    @staticmethod
    def _winner(v1, v2) -> str:
        if v1 > v2:
            return "Method 1"
        elif v2 > v1:
            return "Method 2"
        return "동일"
