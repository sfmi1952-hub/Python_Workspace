import logging
import math
import re
import time
import unicodedata
from dataclasses import dataclass, asdict
from typing import List, Callable, Dict, Optional

import openpyxl
import pandas as pd

from ..openai_core import OpenAICore
from .prompts import SYSTEM_PROMPT, BATCH_EXTRACTION_PROMPT, FALLBACK_BATCH_PROMPT

logger = logging.getLogger(__name__)

ALL_COLUMNS = [
    "상품코드", "담보PMID", "담보명", "담보구분",
    "보기", "보기구분", "납기", "납기구분",
    "시작연령", "시작연령유형", "끝연령", "끝연령유형",
    "예외구분", "납기예외", "보기예외", "예외결과값", "납기보기동일여부",
]

# Fields that come from the layout (input)
INPUT_FIELDS = ["상품코드", "담보PMID", "담보명"]

# Fields that LLM must fill
FILL_FIELDS = [
    "담보구분", "보기", "보기구분", "납기", "납기구분",
    "시작연령", "시작연령유형", "끝연령", "끝연령유형",
    "예외구분", "납기예외", "보기예외", "예외결과값", "납기보기동일여부",
]

BATCH_SIZE = 100


@dataclass
class ExtractionRow:
    상품코드: str = ""
    담보PMID: str = ""
    담보명: str = ""
    담보구분: str = "0"
    보기: str = ""
    보기구분: str = ""
    납기: str = ""
    납기구분: str = ""
    시작연령: str = ""
    시작연령유형: str = "1"
    끝연령: str = ""
    끝연령유형: str = "1"
    예외구분: str = ""
    납기예외: str = ""
    보기예외: str = ""
    예외결과값: str = ""
    납기보기동일여부: str = ""


class LLMExtractor:
    """Use OpenAI GPT to fill coverage parameters from document in batches of 100."""

    def __init__(self, openai_core: OpenAICore, batch_size: int = BATCH_SIZE):
        self.llm = openai_core
        self.batch_size = batch_size

    @staticmethod
    def load_coverage_list(ground_truth_path: str) -> List[Dict[str, str]]:
        """Load ground truth Excel, extract unique (상품코드, 담보PMID, 담보명) rows."""
        df = pd.read_excel(ground_truth_path, engine="openpyxl")

        for col in INPUT_FIELDS:
            if col not in df.columns:
                raise ValueError(
                    f"Ground truth Excel must have column '{col}'. "
                    f"Found columns: {list(df.columns)}"
                )

        subset = df[INPUT_FIELDS].drop_duplicates().reset_index(drop=True)

        coverages = []
        for _, row in subset.iterrows():
            coverages.append({
                "상품코드": str(row["상품코드"]).strip(),
                "담보PMID": str(row["담보PMID"]).strip(),
                "담보명": str(row["담보명"]).strip(),
            })

        return coverages

    # ── Layout 기반 담보 로드 ──────────────────────────────────

    # 비고2(고정) 컬럼에서 추출 제외 대상 값
    EXCLUDED_BIGO2_VALUES = {"2.삭제", "4.부가취소"}

    @staticmethod
    def load_coverage_from_layout(layout_path: str) -> dict:
        """레이아웃 Excel에서 담보 정보 로드 (첫 번째 시트만).

        레이아웃 구조:
          - Row 2, Col C: 상품코드
          - Row 19: 헤더 행
          - Row 20~: 데이터 (C=비고2(고정), D=담보코드, E=표준담보명, F=출력담보명, G=대표담보명)
          - 비고2(고정) 값이 '2.삭제' 또는 '4.부가취소'인 담보는 제외
          - 빈 행 10개 연속 시 종료 (max_row 버그 방지)

        Returns:
            {"coverages": [...], "sheet_name": "ABC000001", "type_label": "1종(일반형)"}
        """
        wb = openpyxl.load_workbook(layout_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]  # 첫 번째 시트만
        sheet_name = ws.title
        logger.info(f"Loading layout from sheet: {sheet_name}")

        # 상품코드 (row 2, col C)
        product_code = str(ws.cell(row=2, column=3).value or "").strip()
        logger.info(f"Product code: {product_code}")

        # 시트명에서 종 정보 추출 (예: "ABC000001" → 시트 부제 또는 row에서 추출)
        # 종 정보는 시트명 뒤 괄호 안, 또는 별도 셀에 있을 수 있음
        type_label = LLMExtractor._extract_type_from_sheet(ws, sheet_name)
        logger.info(f"Type label: {type_label}")

        coverages: List[Dict[str, str]] = []
        empty_count = 0
        excluded_count = 0

        for row_idx in range(20, ws.max_row + 1):
            bigo2 = ws.cell(row=row_idx, column=3).value       # C: 비고2(고정)
            dambo_code = ws.cell(row=row_idx, column=4).value  # D: 담보코드
            std_name = ws.cell(row=row_idx, column=5).value    # E: 표준담보명
            out_name = ws.cell(row=row_idx, column=6).value    # F: 출력담보명
            rep_name = ws.cell(row=row_idx, column=7).value    # G: 대표담보명

            # 빈 행 체크
            if not dambo_code and not std_name:
                empty_count += 1
                if empty_count >= 10:
                    break
                continue
            empty_count = 0

            # 비고2(고정) 필터: '2.삭제' 또는 '4.부가취소' → 제외
            bigo2_str = str(bigo2 or "").strip()
            if bigo2_str in LLMExtractor.EXCLUDED_BIGO2_VALUES:
                excluded_count += 1
                logger.debug(f"Excluded (bigo2={bigo2_str}): {std_name}")
                continue

            coverages.append({
                "상품코드": product_code,
                "담보코드": str(dambo_code or "").strip(),
                "표준담보명": str(std_name or "").strip(),
                "출력담보명": str(out_name or "").strip(),
                "대표담보명": str(rep_name or "").strip(),
            })

        wb.close()
        if excluded_count > 0:
            logger.info(f"Excluded {excluded_count} coverages by 비고2(고정) filter (2.삭제/4.부가취소)")
        logger.info(f"Loaded {len(coverages)} coverages from layout (sheet: {sheet_name})")
        return {
            "coverages": coverages,
            "sheet_name": sheet_name,
            "type_label": type_label,
        }

    @staticmethod
    def _extract_type_from_sheet(ws, sheet_name: str) -> str:
        """시트에서 종 정보(1종, 2종, ...) 추출.

        시트명 패턴: "ABC000001" (순번) → 실제 종 정보는 시트 내용에서 추출
        - 시트명 자체에 "1종", "2종" 등이 포함된 경우
        - 또는 시트의 첫 몇 행에서 종 정보를 탐색
        """
        # 1) 시트명에서 직접 종 정보 찾기
        m = re.search(r"(\d+종(?:\([^)]*\))?)", sheet_name)
        if m:
            return m.group(1)

        # 2) 시트 내 상단 행(1~10)에서 종 정보 찾기
        for row_idx in range(1, 11):
            for col_idx in range(1, 10):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, str):
                    m = re.search(r"(\d+종\s*(?:\([^)]*\))?)", val)
                    if m:
                        return m.group(1).strip()

        # 3) 못 찾으면 시트 순서로 추정 (1번째=1종, 2번째=2종, ...)
        # sheet_name이 ABC000001 형태면 마지막 숫자로 추정
        m = re.search(r"(\d+)$", sheet_name)
        if m:
            num = int(m.group(1))
            return f"{num}종"

        return ""

    @staticmethod
    def filter_markdown_by_type(
        markdown_text: str,
        type_label: str,
        log: Callable = print,
    ) -> str:
        """마크다운에서 특정 종(type)에 해당하는 섹션만 추출.

        문서 구조:
          "가. 1종(일반형)" ~ "나. 2종(...)" 사이가 1종 섹션
          "나. 2종(...)" ~ "다. 3종(...)" 사이가 2종 섹션
          "마. 1종(일반형), 2종(납입면제형) 및 3종(...)" 부터가 공통 섹션

        해당 종의 전용 섹션 + 공통 섹션(해당 종이 포함된 경우) + 주석 반환.

        Args:
            markdown_text: 전체 마크다운 문서
            type_label: "1종(일반형)" 또는 "1종" 등
            log: 로깅 콜백

        Returns:
            필터링된 마크다운 텍스트 (해당 종 관련 표/내용/주석만 포함)
        """
        if not type_label:
            log("[Filter] No type_label provided, using full markdown")
            return markdown_text

        # 종 번호 추출 (예: "1종(일반형)" → "1")
        type_num_match = re.search(r"(\d+)종", type_label)
        if not type_num_match:
            log(f"[Filter] Cannot parse type number from '{type_label}', using full markdown")
            return markdown_text
        type_num = type_num_match.group(1)

        lines = markdown_text.split("\n")
        total_lines = len(lines)

        # ── 섹션 경계 탐색 ──
        # "가.", "나.", "다.", "라.", "마." 등의 한글 번호 + N종 패턴으로 섹션 구분
        # 예: "가. 1종(일반형)", "나. 2종(납입면제형)", "마. 1종(일반형), 2종 ..."
        section_pattern = re.compile(
            r"^[가-힣]\.\s*.*?(\d+종)"
        )

        # 문서 헤더 (보험 종류, 목적 등 — 섹션 4 이전)
        header_end = 0
        for i, line in enumerate(lines):
            if re.search(r"보험기간.*납입기간.*가입나이", line) or \
               re.match(r"\*\*4\.\s*보험기간", line):
                header_end = i
                break

        # 각 종별 섹션 경계 찾기
        sections = []  # [(start_line, end_line, label, type_nums_included)]
        for i, line in enumerate(lines):
            stripped = line.strip()
            m = section_pattern.match(stripped)
            if m:
                # 이 줄에 포함된 모든 종 번호 수집
                type_nums_in_line = re.findall(r"(\d+)종", stripped)
                sections.append({
                    "start": i,
                    "end": None,  # 다음 섹션 시작 전까지
                    "label": stripped,
                    "types": type_nums_in_line,
                })

        # 각 섹션의 end를 다음 섹션 시작으로 설정
        for idx in range(len(sections)):
            if idx + 1 < len(sections):
                sections[idx]["end"] = sections[idx + 1]["start"]
            else:
                # 마지막 종 섹션 → "5. 의무가입" 또는 문서 끝까지
                # "5." 또는 "**5." 패턴으로 끝 찾기
                end_found = False
                for j in range(sections[idx]["start"] + 1, total_lines):
                    if re.match(r"(\*\*)?5\.\s", lines[j].strip()):
                        sections[idx]["end"] = j
                        end_found = True
                        break
                if not end_found:
                    sections[idx]["end"] = total_lines

        if not sections:
            log("[Filter] No sections found, using full markdown")
            return markdown_text

        # ── 해당 종에 관련된 섹션 수집 ──
        collected_parts = []

        # 1) 문서 헤더 (제목, 보험 종류 등)
        if header_end > 0:
            collected_parts.append("\n".join(lines[:header_end]))

        # 2) 섹션 4 제목 행
        if header_end > 0 and header_end < total_lines:
            collected_parts.append(lines[header_end])

        # 3) 해당 종 전용 섹션 (종 번호가 정확히 일치)
        for sec in sections:
            if sec["types"] == [type_num]:
                # 단독 종 섹션
                part = "\n".join(lines[sec["start"]:sec["end"]])
                collected_parts.append(part)
                log(f"[Filter] Including dedicated section: {sec['label'][:60]}... "
                    f"(lines {sec['start']+1}~{sec['end']})")

        # 4) 공통 섹션 (여러 종이 포함되고 해당 종도 포함)
        for sec in sections:
            if len(sec["types"]) > 1 and type_num in sec["types"]:
                part = "\n".join(lines[sec["start"]:sec["end"]])
                collected_parts.append(part)
                log(f"[Filter] Including shared section: {sec['label'][:60]}... "
                    f"(lines {sec['start']+1}~{sec['end']})")

        # 5) 섹션 이후 주석 (주1, 주2, ...) — 해당 종이 언급된 주석만
        # 마지막 수집 섹션 이후부터 다음 큰 제목(5., 6., ...) 전까지의 주석
        last_section_end = max(sec["end"] for sec in sections) if sections else 0
        footnote_lines = []
        for i in range(last_section_end, total_lines):
            stripped = lines[i].strip()
            # "주N)" 패턴의 주석
            if re.match(r"주\d+\)", stripped):
                # 해당 종이 언급된 주석이거나, 종 구분 없는 일반 주석
                if f"{type_num}종" in stripped or not re.search(r"\d+종", stripped):
                    footnote_lines.append(lines[i])
            # 큰 제목("5.", "6." 등)이 나오면 중단
            elif re.match(r"(\*\*)?[5-9]\.\s", stripped) or \
                 re.match(r"(\*\*)\d+\.\s", stripped):
                break

        if footnote_lines:
            collected_parts.append("\n".join(footnote_lines))
            log(f"[Filter] Including {len(footnote_lines)} footnotes")

        filtered = "\n\n".join(collected_parts)
        log(f"[Filter] Filtered markdown: {len(filtered):,} chars "
            f"(from {len(markdown_text):,} chars, "
            f"{len(filtered)/len(markdown_text)*100:.1f}%)")

        return filtered

    @staticmethod
    def normalize_name(name: str) -> str:
        """담보명 정규화: 매칭을 위한 문자열 전처리.

        - 주석 표시(주N)) 제거
        - 전각 괄호 -> 반각
        - 줄바꿈으로 인한 '/ (' 패턴 → '(' 로 병합
        - 특수문자(·∙‧•) 제거
        - 공백/언더바 제거
        - 소문자화
        """
        if not name:
            return ""
        s = name.strip()
        # 주석 표시 제거 (예: 주3), 주4), 주12) 등)
        s = re.sub(r"주\d+\)", "", s)
        # 전각 괄호 -> 반각
        s = s.replace("（", "(").replace("）", ")")
        # 줄바꿈으로 인한 '/ (' 패턴 병합: "입원일당(1일이상) / (요양병원 제외)" → "입원일당(1일이상)(요양병원 제외)"
        s = re.sub(r"\s*/\s*\(", "(", s)
        # 특수 구분점 제거
        s = re.sub(r"[·∙‧•]", "", s)
        # 공백/언더바 제거
        s = re.sub(r"[\s_]+", "", s)
        return s.lower()

    @staticmethod
    def match_coverages_against_markdown(
        raw_coverages: List[Dict[str, str]],
        markdown_text: str,
        log: Callable = print,
    ) -> List[Dict[str, str]]:
        """레이아웃 담보를 markdown 텍스트와 매칭.

        표준담보명 → 출력담보명 → 대표담보명 순으로 정규화된 markdown에서 존재 여부 확인.
        매칭된 담보 반환: {상품코드, 담보PMID(=담보코드), 담보명(=표준담보명)}
        """
        normalized_md = LLMExtractor.normalize_name(markdown_text)
        matched = []
        unmatched_names = []
        unmatched_coverages = []

        for cov in raw_coverages:
            product_code = cov["상품코드"]
            dambo_code = cov["담보코드"]

            # 우선순위: 표준담보명 → 출력담보명 → 대표담보명
            candidates = [
                ("표준담보명", cov.get("표준담보명", "")),
                ("출력담보명", cov.get("출력담보명", "")),
                ("대표담보명", cov.get("대표담보명", "")),
            ]

            found = False
            for name_type, name_val in candidates:
                if not name_val:
                    continue
                norm = LLMExtractor.normalize_name(name_val)
                if norm and norm in normalized_md:
                    std_name = cov.get("표준담보명", "") or name_val
                    out_name = cov.get("출력담보명", "")
                    rep_name = cov.get("대표담보명", "")
                    # 검색명 목록: 표준과 다른 이름들 (LLM이 문서에서 찾을 때 사용)
                    search_names = []
                    for n in [out_name, rep_name]:
                        if n and n != std_name:
                            search_names.append(n)
                    matched.append({
                        "상품코드": product_code,
                        "담보PMID": dambo_code,
                        "담보명": std_name,
                        "검색명": search_names,
                    })
                    found = True
                    break

            if not found:
                display_name = (
                    cov.get("표준담보명", "") or
                    cov.get("출력담보명", "") or
                    dambo_code
                )
                unmatched_names.append(display_name)
                unmatched_coverages.append({
                    "상품코드": product_code,
                    "담보PMID": dambo_code,
                    "담보명": display_name,
                })

        log(f"[Match] {len(matched)}/{len(raw_coverages)} coverages matched in markdown")
        if unmatched_names:
            log(f"[Match] {len(unmatched_names)} unmatched: {unmatched_names[:10]}{'...' if len(unmatched_names) > 10 else ''}")

        return matched, unmatched_coverages

    def extract_with_batches(
        self,
        coverages: List[Dict[str, str]],
        document_markdown: str,
        log: Callable = print,
    ) -> List[ExtractionRow]:
        """Process coverages in batches of BATCH_SIZE through LLM.

        Args:
            coverages: List of dicts with 상품코드/담보PMID/담보명
            document_markdown: Full markdown text of the document (tables included)
            log: Logging callback

        Returns:
            List of ExtractionRow with all fields filled.
        """
        total = len(coverages)
        num_batches = math.ceil(total / self.batch_size)
        log(f"{'─' * 50}")
        log(f"[LLM] >> Extraction Start")
        log(f"[LLM]   Coverages: {total}")
        log(f"[LLM]   Batch size: {self.batch_size}")
        log(f"[LLM]   Total batches: {num_batches}")
        log(f"[LLM]   Document size: {len(document_markdown):,} chars")
        log(f"{'─' * 50}")

        all_rows: List[ExtractionRow] = []
        overall_start = time.time()
        batch_times: List[float] = []

        for batch_idx in range(num_batches):
            start = batch_idx * self.batch_size
            end = min(start + self.batch_size, total)
            batch = coverages[start:end]

            progress_pct = (batch_idx / num_batches) * 100
            eta_str = ""
            if batch_times:
                avg_time = sum(batch_times) / len(batch_times)
                remaining = (num_batches - batch_idx) * avg_time
                eta_min = int(remaining // 60)
                eta_sec = int(remaining % 60)
                eta_str = f" | ETA: {eta_min}m {eta_sec}s"

            log(f"")
            log(f"[LLM] ── Batch {batch_idx + 1}/{num_batches} "
                f"[{progress_pct:.0f}%]{eta_str} ──")
            log(f"[LLM]   Coverages #{start + 1}~#{end} ({len(batch)}건)")
            log(f"[LLM]   First: {batch[0]['담보명']}")
            if len(batch) > 1:
                log(f"[LLM]   Last:  {batch[-1]['담보명']}")

            batch_start = time.time()
            rows = self._process_batch(batch, document_markdown, log)
            batch_elapsed = time.time() - batch_start
            batch_times.append(batch_elapsed)

            log(f"[LLM]   OK {len(rows)} rows extracted in {batch_elapsed:.1f}s")
            all_rows.extend(rows)

        total_elapsed = time.time() - overall_start
        log(f"")
        log(f"{'─' * 50}")
        log(f"[LLM] >> Extraction Complete")
        log(f"[LLM]   Total rows: {len(all_rows)}")
        log(f"[LLM]   Total time: {total_elapsed:.1f}s ({total_elapsed/60:.1f}m)")
        if batch_times:
            log(f"[LLM]   Avg batch time: {sum(batch_times)/len(batch_times):.1f}s")
        log(f"{'─' * 50}")

        return all_rows

    def _process_batch(
        self,
        batch: List[Dict[str, str]],
        document_markdown: str,
        log: Callable,
        _retry_half: bool = True,
    ) -> List[ExtractionRow]:
        """Send one batch of coverages to LLM for parameter extraction.

        토큰 에러(length truncation, empty response) 발생 시
        배치를 절반으로 나눠 재시도합니다.
        """
        # Build coverage list text
        coverage_lines = []
        for i, cov in enumerate(batch, 1):
            line = (
                f"{i}. 상품코드={cov['상품코드']}, "
                f"담보PMID={cov['담보PMID']}, "
                f"담보명={cov['담보명']}"
            )
            # 검색명 추가 (출력담보명/대표담보명 등 문서에서 찾을 별칭)
            search_names = cov.get("검색명", [])
            if search_names:
                line += f", 검색명={' / '.join(search_names)}"
            coverage_lines.append(line)
        coverage_list_text = "\n".join(coverage_lines)

        prompt = BATCH_EXTRACTION_PROMPT.format(
            document_markdown=document_markdown,
            batch_size=len(batch),
            coverage_list=coverage_list_text,
        )

        prompt_chars = len(SYSTEM_PROMPT) + len(prompt)
        log(f"  [LLM] Prompt size: {prompt_chars:,} chars (~{prompt_chars//4:,} tokens est.)")

        # Call LLM
        try:
            log(f"  [LLM] Calling API...")
            t0 = time.time()
            response = self.llm.chat(
                system_prompt=SYSTEM_PROMPT,
                user_prompt=prompt,
                json_mode=True,
                temperature=0.1,
            )
            api_elapsed = time.time() - t0
            log(f"  [LLM] API returned in {api_elapsed:.1f}s, response: {len(response):,} chars")
        except Exception as e:
            error_str = str(e)
            log(f"  [LLM] API Error: {error_str[:200]}")
            # 토큰 관련 에러 시 배치 절반으로 재시도
            if _retry_half and len(batch) > 1 and (
                "length" in error_str.lower() or
                "token" in error_str.lower() or
                "too large" in error_str.lower()
            ):
                return self._split_and_retry(batch, document_markdown, log)
            return self._empty_rows(batch)

        # 빈 응답 (finish_reason=length 등) → 배치 절반으로 재시도
        if not response and _retry_half and len(batch) > 1:
            log(f"  [LLM] WARNING Empty response (likely token limit), splitting...")
            return self._split_and_retry(batch, document_markdown, log)

        # Parse response
        log(f"  [LLM] Parsing JSON response...")
        data = self.llm.parse_json_response(response)
        if data is None:
            log(f"  [LLM] WARNING Parse failed (response preview: {response[:150]}...)")
            # 에러 응답(모델 거부) → 배치 절반으로 재시도
            if _retry_half and len(batch) > 1:
                return self._split_and_retry(batch, document_markdown, log)
            log("  [LLM] Trying fallback prompt...")
            return self._fallback_batch(batch, document_markdown, log)

        rows = self._data_to_rows(data, batch)
        log(f"  [LLM] Parsed {len(data)} JSON items -> {len(rows)} rows")
        return rows

    def _split_and_retry(
        self,
        batch: List[Dict[str, str]],
        document_markdown: str,
        log: Callable,
    ) -> List[ExtractionRow]:
        """배치를 절반으로 나눠 재시도."""
        mid = len(batch) // 2
        first_half = batch[:mid]
        second_half = batch[mid:]
        log(f"  [LLM] Splitting batch: {len(batch)} -> {len(first_half)} + {len(second_half)}")

        log(f"  [LLM]   ── Sub-batch A ({len(first_half)}건) ──")
        rows1 = self._process_batch(first_half, document_markdown, log, _retry_half=True)
        log(f"  [LLM]   Sub-batch A -> {len(rows1)} rows")

        log(f"  [LLM]   ── Sub-batch B ({len(second_half)}건) ──")
        rows2 = self._process_batch(second_half, document_markdown, log, _retry_half=True)
        log(f"  [LLM]   Sub-batch B -> {len(rows2)} rows")

        log(f"  [LLM] Split total: {len(rows1) + len(rows2)} rows")
        return rows1 + rows2

    def _fallback_batch(
        self,
        batch: List[Dict[str, str]],
        document_markdown: str,
        log: Callable,
    ) -> List[ExtractionRow]:
        """Retry with simplified fallback prompt."""
        coverage_lines = []
        for i, cov in enumerate(batch, 1):
            coverage_lines.append(
                f"{i}. 상품코드={cov['상품코드']}, "
                f"담보PMID={cov['담보PMID']}, "
                f"담보명={cov['담보명']}"
            )

        prompt = FALLBACK_BATCH_PROMPT.format(
            document_markdown=document_markdown,
            coverage_list="\n".join(coverage_lines),
        )

        try:
            response = self.llm.chat(
                system_prompt=SYSTEM_PROMPT,
                user_prompt=prompt,
                json_mode=True,
                temperature=0.0,
            )
            data = self.llm.parse_json_response(response)
            if data:
                return self._data_to_rows(data, batch)
        except Exception as e:
            log(f"  [LLM] Fallback also failed: {e}")

        return self._empty_rows(batch)

    def _data_to_rows(
        self, data: list, batch: List[Dict[str, str]]
    ) -> List[ExtractionRow]:
        """Convert LLM JSON output to ExtractionRow objects.

        조합그룹 형식: 각 표 행을 하나의 그룹으로 받아
        그룹별 Cartesian Product로 전개 (잘못된 교차 조합 방지).
        이전 형식(flat 보기_목록/납기_목록, 단일값)도 하위호환 지원.
        """
        from itertools import product as cartesian_product

        rows: List[ExtractionRow] = []
        for item in data:
            if not isinstance(item, dict):
                continue

            base_code = self._clean_val(item.get("상품코드", ""))
            base_pmid = self._clean_val(item.get("담보PMID", ""))
            base_name = self._clean_val(item.get("담보명", ""))
            base_gubun_raw = self._clean_val(item.get("담보구분", "0"), default="0")
            # 담보구분 정제: "0(선택계약)" → "0", "2(기본계약)" → "2" 등
            base_gubun = re.sub(r"[^02]", "", base_gubun_raw)[:1] or "0"
            예외구분 = self._clean_val(item.get("예외구분", ""))
            납기예외 = self._clean_val(item.get("납기예외", ""))
            보기예외 = self._clean_val(item.get("보기예외", ""))
            예외결과값 = self._clean_val(item.get("예외결과값", ""))

            # 조합그룹 형식 (신규 — 권장)
            groups = item.get("조합그룹", [])
            if isinstance(groups, list) and groups:
                for group in groups:
                    if not isinstance(group, dict):
                        continue
                    bogi_list = group.get("보기_목록", [])
                    napgi_list = group.get("납기_목록", [])
                    age_list = group.get("연령_목록", [])

                    if not bogi_list:
                        continue
                    if not napgi_list:
                        napgi_list = [{"값": "", "구분": "", "전기납": False}]
                    if not age_list:
                        age_list = [{"시작": "", "끝": "", "시작유형": "1", "끝유형": "1"}]

                    rows.extend(self._expand_group(
                        bogi_list, napgi_list, age_list,
                        base_code, base_pmid, base_name, base_gubun,
                        예외구분, 납기예외, 보기예외, 예외결과값,
                    ))
                continue

            # 이전 형식 하위호환: flat 보기_목록 (조합그룹 없이)
            bogi_list = item.get("보기_목록", [])
            napgi_list = item.get("납기_목록", [])
            age_list = item.get("연령_목록", [])

            if isinstance(bogi_list, list) and bogi_list:
                if not napgi_list:
                    napgi_list = [{"값": "", "구분": "", "전기납": False}]
                if not age_list:
                    age_list = [{"시작": "", "끝": "", "시작유형": "1", "끝유형": "1"}]

                rows.extend(self._expand_group(
                    bogi_list, napgi_list, age_list,
                    base_code, base_pmid, base_name, base_gubun,
                    예외구분, 납기예외, 보기예외, 예외결과값,
                ))
                continue

            # 최초 형식 하위호환 (단일값)
            rows.append(ExtractionRow(
                상품코드=base_code,
                담보PMID=base_pmid,
                담보명=base_name,
                담보구분=base_gubun,
                보기=self._clean_val(item.get("보기", "")),
                보기구분=self._clean_val(item.get("보기구분", "")),
                납기=self._clean_val(item.get("납기", "")),
                납기구분=self._clean_val(item.get("납기구분", "")),
                시작연령=self._clean_val(item.get("시작연령", "")),
                시작연령유형=self._clean_val(item.get("시작연령유형", "1"), default="1"),
                끝연령=self._clean_val(item.get("끝연령", "")),
                끝연령유형=self._clean_val(item.get("끝연령유형", "1"), default="1"),
                예외구분=예외구분,
                납기예외=납기예외,
                보기예외=보기예외,
                예외결과값=예외결과값,
                납기보기동일여부=self._clean_val(item.get("납기보기동일여부", "")),
            ))
        return rows

    def _expand_group(
        self,
        bogi_list: list,
        napgi_list: list,
        age_list: list,
        base_code: str,
        base_pmid: str,
        base_name: str,
        base_gubun: str,
        예외구분: str,
        납기예외: str,
        보기예외: str,
        예외결과값: str,
    ) -> List[ExtractionRow]:
        """하나의 조합그룹(표 행)에 대해 Cartesian Product 전개."""
        from itertools import product as cartesian_product

        rows: List[ExtractionRow] = []
        for bogi, napgi, age in cartesian_product(bogi_list, napgi_list, age_list):
            bogi_val = self._clean_val(bogi.get("값", "") if isinstance(bogi, dict) else "")
            bogi_gubun = self._clean_val(bogi.get("구분", "") if isinstance(bogi, dict) else "")
            napgi_val = self._clean_val(napgi.get("값", "") if isinstance(napgi, dict) else "")
            napgi_gubun = self._clean_val(napgi.get("구분", "") if isinstance(napgi, dict) else "")
            is_jeongi = napgi.get("전기납", False) if isinstance(napgi, dict) else False

            # 전기납이면 납기값 = 보기값
            if is_jeongi:
                napgi_val = bogi_val

            납기보기동일 = "Y" if is_jeongi else "N"

            시작 = self._clean_val(age.get("시작", "") if isinstance(age, dict) else "", default="")
            끝_raw = self._clean_val(age.get("끝", "") if isinstance(age, dict) else "", default="")
            시작유형 = self._clean_val(age.get("시작유형", "1") if isinstance(age, dict) else "1", default="1")
            끝유형 = self._clean_val(age.get("끝유형", "1") if isinstance(age, dict) else "1", default="1")

            # Min() 수식 계산: Min(A, B-납입기간) → 실제 값 산출
            끝, 끝유형 = self._resolve_min_formula(끝_raw, 끝유형, napgi_val, bogi_val)

            rows.append(ExtractionRow(
                상품코드=base_code,
                담보PMID=base_pmid,
                담보명=base_name,
                담보구분=base_gubun,
                보기=bogi_val,
                보기구분=bogi_gubun,
                납기=napgi_val,
                납기구분=napgi_gubun,
                시작연령=시작,
                시작연령유형=시작유형,
                끝연령=끝,
                끝연령유형=끝유형,
                예외구분=예외구분,
                납기예외=납기예외,
                보기예외=보기예외,
                예외결과값=예외결과값,
                납기보기동일여부=납기보기동일,
            ))
        return rows

    @staticmethod
    def _resolve_min_formula(
        end_val: str, end_type: str, napgi_val: str, bogi_val: str
    ) -> tuple:
        """Min(A, B-납입기간) 수식을 실제 값으로 계산.

        예: Min(15, 30-납입기간), 납기=010 → min(15, 30-10) = min(15, 20) = 15 → "015"
        예: Min(15, 30-납입기간), 납기=025 → min(15, 30-25) = min(15, 5) = 5 → "005"

        Returns:
            (끝연령, 끝연령유형) 튜플. 수식 계산 성공 시 3자리 숫자와 유형 "1" 반환.
        """
        if "Min(" not in end_val and "min(" not in end_val:
            return end_val, end_type

        try:
            napgi_int = int(napgi_val)
        except (ValueError, TypeError):
            return end_val, end_type

        # 패턴1: Min(A, B-납입기간) — 가장 일반적
        m = re.search(r'[Mm]in\(\s*(\d+)\s*,\s*(\d+)\s*-\s*납입기간\s*\)', end_val)
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            result = max(0, min(a, b - napgi_int))
            return f"{result:03d}", "1"

        # 패턴2: Min(A, 보기-납입기간) — 보기값 참조
        m = re.search(r'[Mm]in\(\s*(\d+)\s*,\s*보기\s*-\s*납입기간\s*\)', end_val)
        if m:
            try:
                a = int(m.group(1))
                bogi_int = int(bogi_val)
                result = max(0, min(a, bogi_int - napgi_int))
                return f"{result:03d}", "1"
            except (ValueError, TypeError):
                pass

        return end_val, end_type

    @staticmethod
    def _clean_val(val, default: str = "") -> str:
        """LLM 응답 값 정리: None/nan 처리."""
        s = str(val).strip() if val is not None else ""
        if s.lower() in ("none", "nan", "null", "n/a"):
            return default
        return s if s else default

    def _empty_rows(self, batch: List[Dict[str, str]]) -> List[ExtractionRow]:
        """Return placeholder rows with empty fill fields when LLM fails."""
        rows = []
        for cov in batch:
            rows.append(ExtractionRow(
                상품코드=cov["상품코드"],
                담보PMID=cov["담보PMID"],
                담보명=cov["담보명"],
            ))
        return rows

    @staticmethod
    def export_to_excel(rows: List[ExtractionRow], output_path: str) -> str:
        """Write extraction results to Excel file."""
        data = [asdict(r) for r in rows]
        df = pd.DataFrame(data, columns=ALL_COLUMNS)
        try:
            df.to_excel(output_path, index=False, engine="openpyxl")
        except PermissionError:
            # 파일이 열려있으면 타임스탬프 붙인 파일명으로 저장
            from datetime import datetime as dt
            p = Path(output_path)
            ts = dt.now().strftime("%Y%m%d_%H%M%S")
            alt_path = str(p.parent / f"{p.stem}_{ts}{p.suffix}")
            df.to_excel(alt_path, index=False, engine="openpyxl")
            logger.warning(f"Original file locked, saved to: {alt_path}")
            return alt_path
        return output_path
