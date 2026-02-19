"""
Template Generator Module
Creates the user input Excel template for product and coverage attributes.
"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


class TemplateGenerator:
    """Generates and manages the user input Excel template."""
    
    def __init__(self, template_path: str):
        self.template_path = template_path
        self.wb = None
    
    def create_template(self) -> str:
        """
        Creates a new input template Excel file.
        Returns the path to the created file.
        """
        self.wb = Workbook()
        
        # Create sheets (경로설정 first)
        self._create_path_settings_sheet()
        self._create_product_attributes_sheet()
        self._create_coverage_list_sheet()
        
        # Remove default sheet if exists
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Save the workbook
        os.makedirs(os.path.dirname(self.template_path), exist_ok=True)
        self.wb.save(self.template_path)
        
        return self.template_path
    
    def _create_product_attributes_sheet(self):
        """Creates the 상품속성 (Product Attributes) sheet."""
        ws = self.wb.create_sheet("상품속성")
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["속성명", "값", "설명"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        attributes = [
            ("상품코드", "", "예: P001"),
            ("상품명", "", "예: 무배당 건강보험"),
            ("자동갱신형", 0, "0=아니오, 1=예"),
            ("단체보험", 0, "0=아니오, 1=예"),
            ("모듈형", 0, "0=아니오, 1=예"),
            ("독립특약", 0, "0=아니오, 1=예"),
            ("중증간편", 0, "0=아니오, 1=예"),
            ("0세자녀", 0, "0=아니오, 1=예"),
        ]
        
        for row, (attr_name, default_value, description) in enumerate(attributes, 2):
            ws.cell(row=row, column=1, value=attr_name).border = thin_border
            ws.cell(row=row, column=2, value=default_value).border = thin_border
            ws.cell(row=row, column=3, value=description).border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
    
    def _create_path_settings_sheet(self):
        """Creates the 경로설정 (Path Settings) sheet."""
        ws = self.wb.create_sheet("경로설정")
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["항목", "경로/파일명", "설명"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows - Clear naming: 폴더경로 for directories, 파일경로 for files
        paths = [
            ("약관Base 폴더경로", "", "약관Base 파일들이 있는 폴더 (여러 경로는 ; 로 구분)"),
            ("ExcelPGM 파일경로", "", "ExcelPGM 파일 전체 경로 (예: C:\\PGM\\상품_PGM.xlsm)"),
            ("상품약관 파일경로", "", "상품약관 파일 전체 경로 (예: C:\\약관\\상품약관.docx)"),
            ("출력약관 폴더경로", "", "결과 약관이 저장될 폴더 경로"),
        ]
        
        for row, (item, default_value, desc) in enumerate(paths, 2):
            ws.cell(row=row, column=1, value=item).border = thin_border
            ws.cell(row=row, column=2, value=default_value).border = thin_border
            ws.cell(row=row, column=3, value=desc).border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 50
    
    def _create_coverage_list_sheet(self):
        """Creates the 담보목록 (Coverage List) sheet."""
        ws = self.wb.create_sheet("담보목록")
        
        # Header styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Column categories with colors
        # 사용자 입력: Orange
        user_input_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
        # PGM 참조값: Green  
        pgm_ref_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        # 매핑 결과: Blue
        mapping_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers with category info
        # Format: (header_name, fill_color, width)
        headers = [
            # 순번 (auto)
            ("순번", user_input_fill, 6),
            # 사용자 입력 (7개)
            ("담보코드", user_input_fill, 15),
            ("담보명", user_input_fill, 25),
            ("출력담보명", user_input_fill, 25),
            ("진단확정", user_input_fill, 10),
            ("부모", user_input_fill, 8),
            ("예약가입연령", user_input_fill, 12),
            ("모듈", user_input_fill, 8),
            # PGM 참조값 (3개)
            ("면책", pgm_ref_fill, 8),
            ("감액", pgm_ref_fill, 8),
            ("연장형", pgm_ref_fill, 10),
            # 매핑 결과 (3개)
            ("대표담보코드", mapping_fill, 15),
            ("구분값", mapping_fill, 10),
            ("확인필요", mapping_fill, 10),
        ]
        
        # Create headers
        for col, (header_name, fill, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header_name)
            cell.font = header_font
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Add category row (row 2 as sub-header)
        category_row = [
            "", 
            "사용자입력", "사용자입력", "사용자입력", "사용자입력", "사용자입력", "사용자입력", "사용자입력",
            "PGM참조", "PGM참조", "PGM참조",
            "매핑결과", "매핑결과", "매핑결과"
        ]
        for col, cat in enumerate(category_row, 1):
            cell = ws.cell(row=2, column=col, value=cat)
            cell.font = Font(italic=True, size=9)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Add 100 empty rows for user input (starting from row 3)
        for row in range(3, 103):
            ws.cell(row=row, column=1, value=row - 2).border = thin_border  # 순번
            for col in range(2, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                # Set default values for numeric columns (진단확정, 부모, 예약가입연령, 모듈, 면책, 감액, 연장형)
                if col in [5, 6, 7, 8, 9, 10, 11]:
                    cell.value = 0
        
        # Freeze header rows
        ws.freeze_panes = 'A3'
    
    def open_template(self):
        """Opens the template in Excel for user editing."""
        if not os.path.exists(self.template_path):
            self.create_template()
        
        # Open with default application (Excel)
        os.startfile(self.template_path)
    
    def load_template_data(self) -> dict:
        """
        Loads data from the template Excel file.
        Returns a dictionary with all input data.
        """
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template not found: {self.template_path}")
        
        wb = load_workbook(self.template_path, data_only=True)
        data = {
            'product_attributes': {},
            'path_settings': {},
            'coverage_list': []
        }
        
        # Load 상품속성
        if '상품속성' in wb.sheetnames:
            ws = wb['상품속성']
            for row in range(2, ws.max_row + 1):
                attr_name = ws.cell(row=row, column=1).value
                attr_value = ws.cell(row=row, column=2).value
                if attr_name:
                    data['product_attributes'][attr_name] = attr_value
        
        # Load 경로설정
        if '경로설정' in wb.sheetnames:
            ws = wb['경로설정']
            for row in range(2, ws.max_row + 1):
                item = ws.cell(row=row, column=1).value
                path = ws.cell(row=row, column=2).value
                if item:
                    data['path_settings'][item] = path or ""
        
        # Load 담보목록 (row 1: headers, row 2: category sub-headers, row 3+: data)
        if '담보목록' in wb.sheetnames:
            ws = wb['담보목록']
            headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            
            # Start from row 3 (skip header and category row)
            for row in range(3, ws.max_row + 1):
                row_data = {}
                has_data = False
                
                for col, header in enumerate(headers, 1):
                    value = ws.cell(row=row, column=col).value
                    if header:
                        row_data[header] = value
                        if header != '순번' and value:
                            has_data = True
                
                if has_data:
                    data['coverage_list'].append(row_data)
        
        wb.close()
        return data


def create_default_template(base_path: str) -> str:
    """
    Creates the default template file.
    Returns the path to the created template.
    """
    template_path = os.path.join(base_path, "templates", "입력템플릿.xlsx")
    generator = TemplateGenerator(template_path)
    return generator.create_template()
