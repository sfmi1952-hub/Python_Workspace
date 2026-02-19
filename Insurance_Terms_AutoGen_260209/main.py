"""
Insurance Terms AutoGen - Main Application
Python port of VBA Excel Macro for insurance policy document generation.

Architecture:
- User input via Excel template (상품속성, 경로설정, 담보목록)
- Static mapping via CSV files (담보매핑, 참조)
"""
import sys
import os
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTextEdit, QProgressBar,
    QMessageBox, QGroupBox, QFrame, QTabWidget
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont

from src.template_generator import TemplateGenerator
from src.csv_loader import CSVLoader
from src.mapping_check import MappingChecker
from src.mod_appendix import ModAppendix
from src.print_dambo import PrintDambo
from src.data_loader import DataLoader  # High-performance Pandas loader


# Get the application base path
APP_BASE_PATH = os.path.dirname(os.path.abspath(__file__))


class WorkerThread(QThread):
    """Background worker thread for long-running tasks."""
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)
    finished_signal = pyqtSignal(bool, str)
    
    def __init__(self, task_func):
        super().__init__()
        self.task_func = task_func
    
    def run(self):
        try:
            self.task_func(self.log_signal.emit, self.progress_signal.emit)
            self.finished_signal.emit(True, "완료!")
        except Exception as e:
            self.finished_signal.emit(False, str(e))


class MainWindow(QMainWindow):
    # 약관Base 파일 타입 키워드 (파일명에 포함되면 자동 분류)
    BASE_FILE_TYPES = {
        '상해': '상해',
        '질병': '질병',
        '상해질병': '상해질병',
        '비용배상책임': '비용배상책임',
        '펫': '펫',
        '태아': '태아',
        '부양자': '부양자'
    }
    
    def __init__(self):
        super().__init__()

        self.setWindowTitle("약관 생성 프로그램 (Python Version)")
        self.setGeometry(100, 100, 900, 750)
        self.setMinimumSize(800, 650)

        # Template and CSV paths
        self.template_path = os.path.join(APP_BASE_PATH, "templates", "입력템플릿.xlsx")
        self.data_path = os.path.join(APP_BASE_PATH, "data")
        
        # Data holders
        self.template_generator = TemplateGenerator(self.template_path)
        self.csv_loader = CSVLoader(self.data_path)
        self.template_data = None
        self.data_loader = None
        
        self.worker = None
        
        self.init_ui()
        self.apply_styles()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Title
        title_label = QLabel("📋 약관 생성 프로그램")
        title_label.setFont(QFont("맑은 고딕", 18, QFont.Weight.Bold))
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(title_label)
        
        # Subtitle
        subtitle_label = QLabel("VBA 매크로 → Python 변환 버전 | 입력 템플릿 + CSV 기반")
        subtitle_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        subtitle_label.setStyleSheet("color: #666; font-size: 12px;")
        main_layout.addWidget(subtitle_label)

        # ===== Step 1: 파일 불러오기 =====
        step1_group = QGroupBox("📁 Step 1: 파일 불러오기")
        step1_layout = QVBoxLayout(step1_group)
        
        # Description
        desc_label = QLabel("입력 템플릿(Excel)이 열리고 CSV 매핑 파일이 자동으로 로드됩니다.")
        desc_label.setStyleSheet("color: #555; font-size: 11px; margin-bottom: 5px;")
        step1_layout.addWidget(desc_label)
        
        # Buttons row
        btn_row = QHBoxLayout()
        
        self.btn_load_files = QPushButton("📂 파일 불러오기 (템플릿 열기)")
        self.btn_load_files.setToolTip("입력 템플릿 Excel을 열고 CSV 파일을 로드합니다")
        self.btn_load_files.clicked.connect(self.load_files)
        
        self.btn_reload_template = QPushButton("🔄 템플릿 데이터 새로고침")
        self.btn_reload_template.setToolTip("저장된 템플릿 데이터를 다시 읽어옵니다")
        self.btn_reload_template.clicked.connect(self.reload_template_data)
        self.btn_reload_template.setEnabled(False)
        
        btn_row.addWidget(self.btn_load_files)
        btn_row.addWidget(self.btn_reload_template)
        step1_layout.addLayout(btn_row)
        
        # Status labels
        status_layout = QHBoxLayout()
        self.status_template = QLabel("템플릿: ⏳ 대기중")
        self.status_csv = QLabel("CSV: ⏳ 대기중")
        self.status_template.setStyleSheet("font-size: 11px;")
        self.status_csv.setStyleSheet("font-size: 11px;")
        status_layout.addWidget(self.status_template)
        status_layout.addWidget(self.status_csv)
        step1_layout.addLayout(status_layout)
        
        main_layout.addWidget(step1_group)

        # Tab Widget for different functions
        tab_widget = QTabWidget()
        
        # Tab 1: 담보 매핑
        tab1 = QWidget()
        tab1_layout = QVBoxLayout(tab1)
        
        mapping_group = QGroupBox("🔍 Step 2: 담보 매핑 체크")
        mapping_layout = QVBoxLayout(mapping_group)
        
        mapping_desc = QLabel("담보코드와 대표담보코드 매핑을 확인합니다. 결과는 템플릿의 '대표담보코드', '구분값', '확인필요' 컬럼에 반영됩니다.")
        mapping_desc.setWordWrap(True)
        mapping_desc.setStyleSheet("color: #555; font-size: 11px;")
        mapping_layout.addWidget(mapping_desc)
        
        self.btn_mapping = QPushButton("담보 매핑 체크 실행")
        self.btn_mapping.clicked.connect(self.run_mapping_check)
        self.btn_mapping.setEnabled(False)
        mapping_layout.addWidget(self.btn_mapping)
        
        tab1_layout.addWidget(mapping_group)
        tab_widget.addTab(tab1, "담보 매핑")
        
        # Tab 2: 약관 생성
        tab2 = QWidget()
        tab2_layout = QVBoxLayout(tab2)
        
        generation_group = QGroupBox("📝 Step 3: 약관 생성")
        generation_layout = QVBoxLayout(generation_group)
        
        gen_desc = QLabel("템플릿에 입력된 상품속성, 담보속성 값을 기반으로 약관을 생성합니다.")
        gen_desc.setWordWrap(True)
        gen_desc.setStyleSheet("color: #555; font-size: 11px;")
        generation_layout.addWidget(gen_desc)
        
        self.btn_print = QPushButton("⭐ 약관 출력 (PrintDambo)")
        self.btn_print.setToolTip("PGM 정보를 읽어 약관을 생성합니다")
        self.btn_print.clicked.connect(self.run_print_dambo)
        self.btn_print.setEnabled(False)
        self.btn_print.setStyleSheet("font-weight: bold; font-size: 14px; padding: 15px;")
        
        generation_layout.addWidget(self.btn_print)
        tab2_layout.addWidget(generation_group)
        
        tab_widget.addTab(tab2, "약관 생성")
        
        main_layout.addWidget(tab_widget)

        # Progress Bar
        progress_group = QGroupBox("진행 상황")
        progress_layout = QVBoxLayout(progress_group)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("%p%")
        progress_layout.addWidget(self.progress_bar)
        
        main_layout.addWidget(progress_group)

        # Log Area
        log_group = QGroupBox("📋 실행 로그")
        log_layout = QVBoxLayout(log_group)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 10))
        self.log_text.setMinimumHeight(180)
        log_layout.addWidget(self.log_text)
        
        # Clear log button
        clear_btn = QPushButton("로그 지우기")
        clear_btn.clicked.connect(self.log_text.clear)
        clear_btn.setFixedWidth(100)
        log_layout.addWidget(clear_btn, alignment=Qt.AlignmentFlag.AlignRight)
        
        main_layout.addWidget(log_group)

    def apply_styles(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #ffffff;
            }
            QGroupBox {
                font-weight: bold;
                border: 1px solid #ddd;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
            }
            QPushButton {
                background-color: #4a90d9;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 5px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #357abd;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
            QPushButton:pressed {
                background-color: #2a5d8f;
            }
            QProgressBar {
                border: 1px solid #ddd;
                border-radius: 5px;
                text-align: center;
                height: 25px;
            }
            QProgressBar::chunk {
                background-color: #4CAF50;
                border-radius: 4px;
            }
            QTextEdit {
                border: 1px solid #ddd;
                border-radius: 5px;
                background-color: #fafafa;
            }
            QTabWidget::pane {
                border: 1px solid #ddd;
                border-radius: 5px;
            }
            QTabBar::tab {
                background: #f0f0f0;
                padding: 10px 20px;
                margin-right: 2px;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
            }
            QTabBar::tab:selected {
                background: #4a90d9;
                color: white;
            }
        """)

    def log(self, message):
        self.log_text.append(message)
        self.log_text.verticalScrollBar().setValue(
            self.log_text.verticalScrollBar().maximum()
        )
        QApplication.processEvents()

    def update_progress(self, value):
        self.progress_bar.setValue(value)

    def enable_buttons(self, enabled=True):
        self.btn_reload_template.setEnabled(enabled)
        self.btn_mapping.setEnabled(enabled)
        self.btn_print.setEnabled(enabled)

    def load_files(self):
        """
        Step 1: Load files
        - Creates/opens the input template Excel
        - Loads CSV mapping files
        """
        self.log("=" * 50)
        self.log("📂 Step 1: 파일 불러오기 시작...")
        
        import time
        start_time = time.time()
        
        # 1. Load CSV files first
        self.log("\n📊 CSV 매핑 파일 로드 중...")
        csv_loaded = self.csv_loader.load_all(self.log)
        
        if csv_loaded:
            self.status_csv.setText(f"CSV: ✅ 로드됨 (담보매핑: {len(self.csv_loader.담보매핑_data)}건, 참조: {len(self.csv_loader.참조_data)}건)")
            self.status_csv.setStyleSheet("font-size: 11px; color: green;")
        else:
            self.status_csv.setText("CSV: ⚠️ 일부 파일 누락")
            self.status_csv.setStyleSheet("font-size: 11px; color: orange;")
        
        # 2. Create/open template
        self.log("\n📝 입력 템플릿 Excel 열기...")
        
        try:
            if not os.path.exists(self.template_path):
                self.log("   템플릿 파일이 없습니다. 새로 생성합니다...")
                self.template_generator.create_template()
                self.log(f"   ✅ 템플릿 생성: {self.template_path}")
            
            # Open the template in Excel
            self.template_generator.open_template()
            self.log("   ✅ Excel이 열렸습니다. 값을 입력하고 저장하세요.")
            
            self.status_template.setText(f"템플릿: ✅ 열림")
            self.status_template.setStyleSheet("font-size: 11px; color: green;")
            
            # Enable buttons
            self.enable_buttons(True)
            
            self.log("\n💡 다음 단계:")
            self.log("   1. Excel에서 상품속성, 경로설정, 담보목록 입력")
            self.log("   2. Excel 저장 (Ctrl+S)")
            self.log("   3. '템플릿 데이터 새로고침' 클릭")
            self.log("   4. '담보 매핑 체크' 또는 '약관 출력' 실행")
            
            end_time = time.time()
            elapsed_time = end_time - start_time
            self.log(f"⏱️ Step 1 소요 시간: {elapsed_time:.2f}초")
            
        except Exception as e:
            self.log(f"❌ 오류: {e}")
            QMessageBox.critical(self, "오류", str(e))

    def reload_template_data(self):
        """Reloads data from the saved template."""
        self.log("\n🔄 템플릿 데이터 새로고침...")
        
        import time
        start_time = time.time()
        
        try:
            self.template_data = self.template_generator.load_template_data()
            
            product_attrs = self.template_data.get('product_attributes', {})
            path_settings = self.template_data.get('path_settings', {})
            coverage_count = len(self.template_data.get('coverage_list', []))
            
            self.log(f"   상품코드: {product_attrs.get('상품코드', '미입력')}")
            self.log(f"   상품명: {product_attrs.get('상품명', '미입력')}")
            self.log(f"   담보 수: {coverage_count}건")
            
            # Show path settings
            self.log("\n   📁 경로 설정:")
            for key, value in path_settings.items():
                self.log(f"      {key}: {value or '(미입력)'}")
            
            if coverage_count > 0:
                self.log("\n   ✅ 템플릿 데이터 로드 완료!")
            else:
                self.log("   ⚠️ 담보목록이 비어있습니다. Excel에서 입력해주세요.")
            
            end_time = time.time()
            elapsed_time = end_time - start_time
            self.log(f"⏱️ 템플릿 새로고침 소요 시간: {elapsed_time:.2f}초")
                
        except FileNotFoundError:
            self.log("   ❌ 템플릿 파일을 찾을 수 없습니다.")
        except Exception as e:
            self.log(f"   ❌ 오류: {e}")

    def run_mapping_check(self):
        """
        Runs the coverage mapping check.
        - Loads PGM data for 면책/감액/연장형
        - Performs CSV mapping for 대표담보코드/구분값
        - Saves results back to template Excel
        """
        self.log("\n🔍 담보 매핑 체크 시작...")
        
        # Reload template data first
        self.reload_template_data()
        
        if not self.template_data:
            self.log("❌ 템플릿 데이터가 없습니다.")
            return
        
        path_settings = self.template_data.get('path_settings', {})
        pgm_path = path_settings.get('ExcelPGM 파일경로', '')
        
        if not pgm_path or not os.path.exists(pgm_path):
            self.log("❌ ExcelPGM 파일경로가 입력되지 않았거나 파일이 없습니다.")
            self.log("   경로설정 탭에서 'ExcelPGM 파일경로'를 확인해주세요.")
            return
            
        coverage_list = self.template_data.get('coverage_list', [])
        if not coverage_list:
            self.log("❌ 담보목록이 비어있습니다. 템플릿을 먼저 입력해주세요.")
            return
        
        try:
            import time
            start_time = time.time()
            
            # Load PGM data for 면책/감액/연장형
            self.log("\n📊 PGM 데이터 로드 중...")
            self.data_loader = DataLoader()
            self.data_loader.load_pgm_excel(pgm_path, log_callback=self.log)
            
            # PGM 데이터 확인
            if self.data_loader.df_보장배수 is None or self.data_loader.df_보장배수.empty:
                self.log("   ⚠️ 보장배수 데이터가 없습니다. (면책/감액/연장형 기본값 0 적용)")
            else:
                self.log("   ✅ PGM 데이터 로드 완료")
            
            updated_count = 0
            error_count = 0
            
            self.log("\n📋 담보 매핑 체크 진행...")
            
            for coverage in coverage_list:
                담보코드 = str(coverage.get('담보코드', '')).strip()
                if not 담보코드:
                    continue
                
                # 1. CSV mapping for 대표담보코드, 구분값
                mapping_result = self.csv_loader.find_대표담보코드(담보코드)
                
                if mapping_result:
                    coverage['대표담보코드'] = mapping_result.get('대표담보코드', '')
                    coverage['구분값'] = mapping_result.get('구분값', '')
                    coverage['확인필요'] = ''
                    updated_count += 1
                else:
                    coverage['대표담보코드'] = ''
                    coverage['구분값'] = ''
                    coverage['확인필요'] = '찾기에러'
                    error_count += 1
                
                # 2. PGM lookup for 면책/감액/연장형 (VBA 로직과 동일)
                # VBA Logic: 보장구조 → 보장배수 lookup → sum 면책기간/감액기간 → if sum>0 then 1 else 0
                면책_flag = 0
                감액_flag = 0
                연장형_flag = 0
                
                try:
                    보장구조_df = self.data_loader.df_보장구조
                    보장배수_df = self.data_loader.df_보장배수
                    main_df = self.data_loader.df_pgm_main
                    
                    if 보장구조_df is not None and not 보장구조_df.empty:
                        # 담보코드로 보장구조 시트에서 보장배수 행 인덱스 찾기
                        # VBA: 보장구조_Lookup_Key = ProductCode & "_" & DAMBO
                        # 첫 번째 컬럼에서 담보코드 포함 검색
                        col0_values = 보장구조_df.iloc[:, 0].astype(str).str.strip()
                        matches = 보장구조_df[col0_values.str.contains(담보코드[-7:], na=False)]
                        
                        if not matches.empty:
                            # 보장배수 컬럼 찾기
                            보장배수_col_idx = None
                            면책기간_col_idx = None
                            감액기간_col_idx = None
                            
                            col_names = [str(c) if c else '' for c in 보장배수_df.columns]
                            
                            # 보장배수 시트에서 컬럼 위치 찾기
                            for idx, col_name in enumerate(col_names):
                                if '면책기간' in col_name:
                                    면책기간_col_idx = idx
                                elif '감액기간' in col_name and 감액기간_col_idx is None:
                                    감액기간_col_idx = idx
                            
                            # 보장배수 시트에서 담보코드로 직접 검색
                            sum_면책 = 0
                            sum_감액 = 0
                            
                            for col_idx in range(min(5, len(보장배수_df.columns))):
                                col_values = 보장배수_df.iloc[:, col_idx].astype(str).str.strip()
                                담보_matches = 보장배수_df[col_values == 담보코드]
                                
                                if not 담보_matches.empty:
                                    for _, row in 담보_matches.iterrows():
                                        if 면책기간_col_idx is not None:
                                            val = row.iloc[면책기간_col_idx]
                                            try:
                                                sum_면책 += float(val) if val and str(val) != 'nan' else 0
                                            except: pass
                                        if 감액기간_col_idx is not None:
                                            val = row.iloc[감액기간_col_idx]
                                            try:
                                                sum_감액 += float(val) if val and str(val) != 'nan' else 0
                                            except: pass
                                    break
                            
                            # VBA: If sum_면책 > 0 Then 1 Else 0
                            면책_flag = 1 if sum_면책 > 0 else 0
                            감액_flag = 1 if sum_감액 > 0 else 0
                    
                    # 연장형은 Main 시트의 보험기간연장형 컬럼에서 가져옴
                    if main_df is not None and not main_df.empty:
                        main_col_names = [str(c) if c else '' for c in main_df.columns]
                        연장형_col_idx = None
                        담보코드_col_idx = None
                        
                        for idx, col_name in enumerate(main_col_names):
                            if '보험기간연장형' in col_name:
                                연장형_col_idx = idx
                            if idx < 5:  # 담보코드는 앞쪽 컬럼에 위치
                                담보코드_col_idx = idx
                        
                        if 연장형_col_idx is not None:
                            for col_idx in range(min(5, len(main_df.columns))):
                                col_values = main_df.iloc[:, col_idx].astype(str).str.strip()
                                main_matches = main_df[col_values == 담보코드]
                                if not main_matches.empty:
                                    val = main_matches.iloc[0, 연장형_col_idx]
                                    try:
                                        연장형_flag = 1 if val and float(val) == 1 else 0
                                    except: pass
                                    break
                                    
                except Exception as e:
                    pass  # 에러 시 기본값 0 유지
                
                coverage['면책'] = 면책_flag
                coverage['감액'] = 감액_flag
                coverage['연장형'] = 연장형_flag
                
                status = coverage.get('대표담보코드', '') or '찾기에러'
                self.log(f"   {담보코드} → {status}")
            
            # 3. Save results back to template Excel
            self.log("\n💾 매핑 결과를 템플릿에 저장 중...")
            self._save_mapping_results_to_template(coverage_list)
            
            end_time = time.time()
            elapsed_time = end_time - start_time
            self.log(f"\n✅ 매핑 완료: {updated_count}건 성공, {error_count}건 오류")
            self.log(f"⏱️ Step 2 소요 시간: {elapsed_time:.2f}초")
            
            if error_count > 0:
                self.log("⚠️ 매핑되지 않은 담보는 '확인필요' 컬럼에 '찾기에러'로 표시됩니다.")
                self.log("   data/담보매핑.csv 파일에 해당 담보코드를 추가해주세요.")
            
            self.log("\n💡 템플릿 파일이 업데이트되었습니다. Excel에서 확인하세요.")
                
        except Exception as e:
            self.log(f"❌ 매핑 체크 오류: {e}")
            import traceback
            traceback.print_exc()
    
    def _save_mapping_results_to_template(self, coverage_list):
        """Save mapping results back to template Excel file."""
        from openpyxl import load_workbook
        import shutil
        
        # Check if file is locked (Excel has it open)
        try:
            # Try to open file for writing to check if locked
            with open(self.template_path, 'a'):
                pass
        except PermissionError:
            self.log("   ⚠️ 템플릿 파일이 Excel에서 열려있습니다.")
            self.log("   Excel을 닫고 다시 시도하거나, 수동으로 값을 입력하세요.")
            
            # Show results in log so user can manually enter them
            self.log("\n📋 매핑 결과 (수동 입력용):")
            for i, coverage in enumerate(coverage_list[:10]):  # Show first 10
                담보코드 = coverage.get('담보코드', '')
                대표담보코드 = coverage.get('대표담보코드', '')
                구분값 = coverage.get('구분값', '')
                면책 = coverage.get('면책', 0)
                감액 = coverage.get('감액', 0)
                연장형 = coverage.get('연장형', 0)
                self.log(f"   {담보코드}: 대표={대표담보코드}, 구분={구분값}, 면책={면책}, 감액={감액}, 연장형={연장형}")
            if len(coverage_list) > 10:
                self.log(f"   ... 외 {len(coverage_list) - 10}건")
            return
        
        try:
            wb = load_workbook(self.template_path)
            ws = wb['담보목록']
            
            # Column mapping (1-indexed, new structure)
            # Row 1: headers, Row 2: category, Row 3+: data
            col_map = {
                '면책': 9,        # I
                '감액': 10,       # J
                '연장형': 11,     # K
                '대표담보코드': 12, # L
                '구분값': 13,     # M
                '확인필요': 14,   # N
            }
            
            for idx, coverage in enumerate(coverage_list):
                row = idx + 3  # Data starts from row 3
                
                for field, col in col_map.items():
                    value = coverage.get(field, '')
                    ws.cell(row=row, column=col, value=value)
            
            wb.save(self.template_path)
            wb.close()
            self.log(f"   ✅ 템플릿 저장 완료: {self.template_path}")
            
        except PermissionError:
            self.log("   ❌ 파일 저장 실패: Excel에서 파일을 닫아주세요.")
        except Exception as e:
            self.log(f"   ❌ 템플릿 저장 오류: {e}")


    def run_print_dambo(self):
        """
        Runs the terms generation using template path settings.
        Uses: 약관Base 폴더경로, ExcelPGM 파일경로, 상품약관 파일경로, 출력약관 폴더경로
        """
        self.log("\n⭐ 약관 생성 시작 (최적화 버전)...")
        self.progress_bar.setValue(0)
        
        # Reload template data first
        self.reload_template_data()
        
        if not self.template_data:
            self.log("❌ 템플릿 데이터가 없습니다.")
            return
        
        path_settings = self.template_data.get('path_settings', {})
        
        # Get paths from template (new naming: 폴더경로/파일경로)
        약관Base폴더 = path_settings.get('약관Base 폴더경로', '')
        pgm_path = path_settings.get('ExcelPGM 파일경로', '')
        상품약관경로 = path_settings.get('상품약관 파일경로', '')
        출력약관폴더 = path_settings.get('출력약관 폴더경로', '')
        
        # Validate required paths
        if not 약관Base폴더:
            self.log("❌ 약관Base 폴더경로가 입력되지 않았습니다.")
            return
        
        if not pgm_path:
            self.log("❌ ExcelPGM 파일경로가 입력되지 않았습니다.")
            return
        
        if not 상품약관경로:
            self.log("❌ 상품약관 파일경로가 입력되지 않았습니다.")
            return
        
        if not 출력약관폴더:
            self.log("❌ 출력약관 폴더경로가 입력되지 않았습니다.")
            return
        
        # Validate files exist
        if not os.path.exists(pgm_path):
            self.log(f"❌ ExcelPGM 파일을 찾을 수 없습니다: {pgm_path}")
            return
        
        if not os.path.exists(상품약관경로):
            self.log(f"❌ 상품약관 파일을 찾을 수 없습니다: {상품약관경로}")
            return
        
        # Process 약관Base 폴더 (supports multiple paths separated by ;)
        base_dirs = [d.strip() for d in 약관Base폴더.split(';') if d.strip()]
        base_files = {}
        
        for base_dir in base_dirs:
            if not os.path.exists(base_dir):
                self.log(f"   ⚠️ 약관Base 폴더를 찾을 수 없습니다: {base_dir}")
                continue
            
            # Find Word files in directory and classify (skip temp files)
            for filename in os.listdir(base_dir):
                # Skip Word temp files (created when document is open)
                if filename.startswith('~$'):
                    continue
                if filename.lower().endswith(('.docx', '.doc')):
                    file_path = os.path.join(base_dir, filename)
                    file_type = self._classify_base_file(filename)
                    base_files[file_type] = file_path
                    self.log(f"   ✅ [{file_type}] {filename}")
        
        if not base_files:
            self.log("❌ 약관Base 폴더에서 Word 파일을 찾을 수 없습니다.")
            return
        
        self.log(f"\n   약관Base 파일: {len(base_files)}개 타입 로드됨")
        self.log(f"   PGM 파일: {os.path.basename(pgm_path)}")
        self.log(f"   상품약관 파일: {os.path.basename(상품약관경로)}")
        self.log(f"   출력 폴더: {출력약관폴더}")
        
        # ============ Use DataLoader ============
        import time
        start_time = time.time()
        
        try:
            self.log("\n📊 데이터 로딩 중 (Pandas 최적화)...")
            
            # Create DataLoader and load PGM data
            self.data_loader = DataLoader()
            self.data_loader.load_pgm_excel(pgm_path, log_callback=self.log)
            
            # Set file paths in data_loader
            self.data_loader.base_files = base_files
            self.data_loader.product_doc_file = 상품약관경로
            self.data_loader.csv_loader = self.csv_loader
            self.data_loader.template_data = self.template_data
            
            load_time = time.time() - start_time
            self.log(f"   ⚡ 데이터 로드 완료: {load_time:.2f}초")
            
            # Generate output file path
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            original_name = os.path.splitext(os.path.basename(상품약관경로))[0]
            
            # Ensure output directory exists
            os.makedirs(출력약관폴더, exist_ok=True)
            output_file = os.path.join(출력약관폴더, f"{original_name}_{timestamp}.docx")
            
            self.data_loader.output_file = output_file
            self.log(f"   📤 출력 파일: {output_file}")
            
            # Run PrintDambo with optimized data
            self.log("\n📝 약관 생성 중...")
            
            def run_task(log_cb, progress_cb):
                import time
                task_start_time = time.time()
                
                printer = PrintDambo(self.data_loader)
                printer.base_files = base_files
                printer.csv_loader = self.csv_loader
                printer.execute(log_callback=log_cb, progress_callback=progress_cb)
                
                task_end_time = time.time()
                elapsed_time = task_end_time - task_start_time
                log_cb(f"\n⏱️ Step 3 소요 시간: {elapsed_time:.2f}초 ({int(elapsed_time // 60)}분 {int(elapsed_time % 60)}초)")
            
            # Run in background thread
            self.worker = WorkerThread(run_task)
            self.worker.log_signal.connect(self.log)
            self.worker.progress_signal.connect(self.update_progress)
            self.worker.finished_signal.connect(self._on_print_finished)
            self.worker.start()
            
            # Disable buttons while running
            self.enable_buttons(False)
            
        except Exception as e:
            self.log(f"❌ 오류: {e}")
            import traceback
            traceback.print_exc()
    
    def _classify_base_file(self, filename: str) -> str:
        """Classify base file type based on filename keywords."""
        filename_lower = filename.lower()
        
        # 우선순위: 복합 키워드 먼저 체크
        if '상해질병' in filename_lower:
            return '상해질병'
        elif '비용배상책임' in filename_lower:
            return '비용배상책임'
        elif '상해' in filename_lower:
            return '상해'
        elif '질병' in filename_lower:
            return '질병'
        elif '펫' in filename_lower:
            return '펫'
        elif '태아' in filename_lower:
            return '태아'
        elif '부양자' in filename_lower:
            return '부양자'
        else:
            return '기타'
    
    def _on_print_finished(self, success, message):
        """Callback when print job finishes."""
        self.enable_buttons(True)
        if success:
            self.log(f"\n✅ {message}")
            self.progress_bar.setValue(100)
            
            # Show output file location
            if hasattr(self.data_loader, 'output_file') and self.data_loader.output_file:
                self.log(f"📁 출력 파일: {self.data_loader.output_file}")
        else:
            self.log(f"\n❌ 오류: {message}")

    def closeEvent(self, event):
        event.accept()


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
